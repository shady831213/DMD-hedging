import openpyxl as exl
import numpy as np
import matplotlib.pyplot as plt
import scipy.stats as ss


################################################
# class
################################################
# currence data class
class currence:
    def __init__(self, name):
        self._name = name

    @property
    def nOfSamples(self):
        return self._nOfSamples

    @nOfSamples.setter
    def nOfSamples(self, value):
        self._nOfSamples = value

    @property
    def mean(self):
        return self._mean

    @mean.setter
    def mean(self, value):
        self._mean = value

    @property
    def std(self):
        return self._std

    @std.setter
    def std(self, value):
        self._std = value

    @property
    def curValue(self):
        return self._curValue

    @curValue.setter
    def curValue(self, value):
        self._curValue = value

    @property
    def revenue(self):
        return self._revenue

    @revenue.setter
    def revenue(self, value):
        self._revenue = value

    @property
    def k(self):
        return self._k

    @k.setter
    def k(self, value):
        self._k = value

    @property
    def c(self):
        return self._c

    @c.setter
    def c(self, value):
        self._c = value

    @property
    def nOfOpt(self):
        return self._nOfOpt

    @nOfOpt.setter
    def nOfOpt(self, value):
        self._nOfOpt = value

    @property
    def samples(self):
        return self._samples

    @samples.setter
    def samples(self, value):
        self._samples = value

    def exRate(self):
        return self.curValue * (1.0 + self.samples)

    def netPayoff(self):
        return np.maximum(self.k - self.exRate(), 0) - self.c

    def unhedgedRevenue(self):
        return self.revenue * self.exRate()

    def hedgedRevenue(self):
        return self.unhedgedRevenue() + self.nOfOpt * self.netPayoff()

    def drawHist(self, path):
        plt.hist(self.samples)
        plt.title(self._name)
        plt.xlabel('change rate')
        plt.ylabel('frequency')
        plt.savefig(path + '/hist_' + self._name)
        plt.close('all')

    def dump(self):
        for name, value in vars(self).items():
            print ('%s=' % (name))
            print (value)


################################################
# excel access
################################################
# read datas sheet
def readDatas(book, dm, bp, covArray):
    sheet = book.get_sheet_by_name('Datas')
    cells = sheet['B2':'C7']
    dm.nOfSamples, bp.nOfSamples = map(lambda x: x.value, cells.next())
    dm.mean, bp.mean = map(lambda x: x.value, cells.next())
    dm.std, bp.std = map(lambda x: x.value, cells.next())
    covArray[1, 0], covArray[0, 1] = map(lambda x: x.value * dm.std * bp.std, cells.next())
    dm.curValue, bp.curValue = map(lambda x: x.value, cells.next())
    dm.revenue, bp.revenue = map(lambda x: x.value, cells.next())
    covArray[0, 0] = dm.std ** 2
    covArray[1, 1] = bp.std ** 2


# read number of put options
def readNOfOpt(book, nOfOptListDM, nOfOptListBP):
    sheet = book.get_sheet_by_name('NOFOpt')
    cells = sheet['B2':'C5']
    for cell in cells:
        value0, value1 = map(lambda x: x.value, cell)
        nOfOptListDM.append(value0)
        nOfOptListBP.append(value1)


# read number of put options
def readKAndC(book, kAndCListDM, kAndCListBP):
    sheet = book.get_sheet_by_name('KAndC')
    cells = sheet['A3':'D11']
    for cell in cells:
        k0, c0, k1, c1 = map(lambda x: x.value, cell)
        kAndCListDM.append((k0, c0))
        kAndCListBP.append((k1, c1))


# output 2d table
def write2DTable(sheet, title, dmKC, bpKC, values, row_offset=1):
    titleCell = sheet.cell('A%d' % (row_offset))
    dmTitleCell = sheet.cell('A%d' % (row_offset + 3))
    bpTitleCell = sheet.cell('D%d' % (row_offset))
    dmIndexCells = sheet['B%d' % (row_offset + 3):'C%d' % (row_offset + 12)]
    bpIndexCells = sheet['D%d' % (row_offset + 1):'M%d' % (row_offset + 2)]
    valuesCells = sheet['E%d' % (row_offset + 4):'M%d' % (row_offset + 12)]
    # fill title
    titleCell.value = title
    dmTitleCell.value = 'DM'
    bpTitleCell.value = 'BP'
    sheet.merge_cells('A%d:C%d' % (row_offset, row_offset + 2))
    # fill Index
    dmIndexTitleCellK, dmIndexTitleCellC = dmIndexCells.next()
    dmIndexTitleCellK.value = 'k'
    dmIndexTitleCellC.value = 'c'
    for k, c in dmKC:
        dmIndexCellK, dmIndexCellC = dmIndexCells.next()
        dmIndexCellK.value = k
        dmIndexCellC.value = c
    sheet.merge_cells('A%d:A%d' % (row_offset + 3, row_offset + 12))

    bpIndexLine1 = bpIndexCells.next()
    bpIndexTitleCellK = bpIndexLine1[0]
    bpIndexTitleCellK.value = 'k'
    bpIndexLine2 = bpIndexCells.next()
    bpIndexTitleCellC = bpIndexLine2[0]
    bpIndexTitleCellC.value = 'c'
    for i in range(1, len(bpIndexLine1)):
        bpIndexLine1[i].value = bpKC[i - 1][0]
        bpIndexLine2[i].value = bpKC[i - 1][1]
    sheet.merge_cells('D%d:M%d' % (row_offset, row_offset))

    valuesIter = iter(values)
    # fill values
    for row in valuesCells:
        for cell in row:
            cell.value = valuesIter.next()


################################################
# computing
################################################
# generate samples
def generateSamples(dm, bp, covArray):
    sampleArray = np.random.multivariate_normal(np.array([dm.mean, bp.mean]), covArray, dm.nOfSamples)
    dm.samples = sampleArray[:, 0]
    bp.samples = sampleArray[:, 1]


# computing features
def compFeatures(datas, confidentLevel):
    mean = np.average(datas)
    std = np.std(a=datas, ddof=1)
    limitation = ss.norm.ppf(q=confidentLevel + (1 - confidentLevel) / 2) * (std / np.sqrt(datas.size))
    return mean, std, (mean - limitation, mean + limitation)


# computing propotion
def compProp(datas, confidentLevel, lowTh):
    prop = float(datas[datas > lowTh].size) / float(datas.size)
    std = np.sqrt(prop * (1 - prop) / datas.size)
    limitation = ss.norm.ppf(q=confidentLevel + (1 - confidentLevel) / 2) * std
    # return prop, std, (prop - limitation, prop + limitation)
    return prop, (prop - limitation, prop + limitation)


# main
if __name__ == '__main__':
    ################################################
    # init excel
    ################################################
    # get work book
    inputBook = exl.load_workbook('datas/datas.xlsx')
    # output book
    outputBook = exl.Workbook()
    # outputWriter = exl.writer.excel.ExcelWriter(workbook=outputBook)

    ##################################################
    # read excel
    ##################################################
    '''variables
        #dm
        #bp
        #covArray
        #nOfOptListDM
        #nOfOptListBP
        #kAndCListDM
        #kAndCListBP
    '''
    # dm and bp object
    dm = currence('dm')
    bp = currence('bp')
    # cov array
    covArray = np.ones((2, 2))
    # list of n of opts
    nOfOptListDM = []
    nOfOptListBP = []
    # list of (k,c)
    kAndCListDM = []
    kAndCListBP = []

    # read datas sheet
    readDatas(inputBook, dm, bp, covArray)
    # read n of opts sheet
    readNOfOpt(inputBook, nOfOptListDM, nOfOptListBP)
    # read k and c sheet
    readKAndC(inputBook, kAndCListDM, kAndCListBP)

    ##################################################
    # simulation
    ##################################################
    '''variables
        #corrVerif
    '''
    # output work book

    # generate samples
    generateSamples(dm, bp, covArray)
    # draw hist
    dm.drawHist('output')
    bp.drawHist('output')
    # verify dist
    covArrayVerify = np.cov(np.vstack((dm.samples, bp.samples)))
    corrVerif = covArrayVerify[1, 0] / (np.sqrt(covArrayVerify[0, 0]) * np.sqrt(covArrayVerify[1, 1]))

    for iOfN in range(len(nOfOptListDM)):
        ##################################################
        # computing
        ##################################################
        '''variables
             #hedgedRevenueDM
             #hedgedRevenueBP
             #hedgedRevenue
             #hedgedRevenueMean
             #hedgedRevenueStd
             #hedgedRevenueInterval
             #hedgedRevenueProp
             #hedgedRevenuePropInterval
             #unhedgedRevenueDM
             #unhedgedRevenueBP
             #unhedgedRevenue
             #unhedgedRevenueMean
             #unhedgedRevenueStd
             #unhedgedRevenueInterval
             #unhedgedRevenueProp
             #unhedgedRevenuePropInterval
         '''
        hedgedRevenueDMList = []
        hedgedRevenueBPList = []
        hedgedRevenueList = []
        hedgedRevenueMeanList = []
        hedgedRevenueStdList = []
        hedgedRevenueIntervalList = []
        hedgedRevenuePropList = []
        hedgedRevenuePropIntervalList = []
        unhedgedRevenueDMList = []
        unhedgedRevenueBPList = []
        unhedgedRevenueList = []
        unhedgedRevenueMeanList = []
        unhedgedRevenueStdList = []
        unhedgedRevenueIntervalList = []
        unhedgedRevenuePropList = []
        unhedgedRevenuePropIntervalList = []

        confidentLevel = 0.95
        lowTh = 706.0

        # computing

        for i in range(len(kAndCListDM)):
            for j in range(len(kAndCListBP)):
                dm.nOfOpt = nOfOptListDM[iOfN]
                bp.nOfOpt = nOfOptListBP[iOfN]
                dm.k, dm.c = kAndCListDM[i]
                bp.k, bp.c = kAndCListBP[j]
                # hedged revenue
                hedgedRevenueDM = dm.hedgedRevenue()
                hedgedRevenueBP = bp.hedgedRevenue()
                hedgedRevenue = hedgedRevenueDM + hedgedRevenueBP
                hedgedRevenueDMList.append(hedgedRevenueDM)
                hedgedRevenueBPList.append(hedgedRevenueBP)
                hedgedRevenueList.append(hedgedRevenue)
                # mean,std, confident interval
                hedgedRevenueMean, hedgedRevenueStd, hedgedRevenueInterval = compFeatures(hedgedRevenue, confidentLevel)
                hedgedRevenueMeanList.append(hedgedRevenueMean)
                hedgedRevenueStdList.append(hedgedRevenueStd)
                hedgedRevenueIntervalList.append(hedgedRevenueInterval)
                # prop, prop confident interval
                hedgedRevenueProp, hedgedRevenuePropInterval = compProp(hedgedRevenue, confidentLevel, lowTh)
                hedgedRevenuePropList.append(hedgedRevenueProp)
                hedgedRevenuePropIntervalList.append(hedgedRevenuePropInterval)

                # unhedged revenue
                unhedgedRevenueDM = dm.unhedgedRevenue()
                unhedgedRevenueBP = bp.unhedgedRevenue()
                unhedgedRevenue = unhedgedRevenueDM + unhedgedRevenueBP
                unhedgedRevenueDMList.append(unhedgedRevenueDM)
                unhedgedRevenueBPList.append(unhedgedRevenueBP)
                unhedgedRevenueList.append(unhedgedRevenue)
                # mean,std, confident interval
                unhedgedRevenueMean, unhedgedRevenueStd, unhedgedRevenueInterval = compFeatures(unhedgedRevenue,
                                                                                                confidentLevel)
                unhedgedRevenueMeanList.append(unhedgedRevenueMean)
                unhedgedRevenueStdList.append(unhedgedRevenueStd)
                unhedgedRevenueIntervalList.append(unhedgedRevenueInterval)
                # prop, prop confident interval
                unhedgedRevenueProp, unhedgedRevenuePropInterval = compProp(unhedgedRevenue, confidentLevel, lowTh)
                unhedgedRevenuePropList.append(unhedgedRevenueProp)
                unhedgedRevenuePropIntervalList.append(unhedgedRevenuePropInterval)

                # draw one opt hist
                if iOfN == 0 and i == 0 and j == 0:
                    plt.hist(hedgedRevenue)
                    plt.title('hedgedRevenue\n n=%d kDM=%f cDM=%f kBP=%f cBP=%f' % (dm.nOfOpt, dm.k, dm.c, bp.k, bp.c))
                    plt.xlabel('Revenue')
                    plt.ylabel('frequency')
                    plt.savefig('output/hist_hedgedRevenue_%d_dmOPT%d_bpOPT%d' % (dm.nOfOpt, i, j))
                    plt.close('all')

                    plt.hist(unhedgedRevenue)
                    plt.title(
                        'unhedgedRevenue\n n=%d kDM=%f cDM=%f kBP=%f cBP=%f' % (dm.nOfOpt, dm.k, dm.c, bp.k, bp.c))
                    plt.xlabel('Revenue')
                    plt.ylabel('frequency')
                    plt.savefig('output/hist_unhedgedRevenue_%d_dmOPT%d_bpOPT%d' % (dm.nOfOpt, i, j))
                    plt.close('all')

        ##################################################
        # save result
        ##################################################

        # sheet as simulation
        if iOfN == 0:
            simSheet = outputBook.worksheets[0]
            simSheet.title = 'SIMULATION'
            # output corrVerif
            simSheet.cell('A1').value = 'corrVerif'
            simSheet.cell('B1').value = corrVerif
            # output simulation result
            simSheet.cell('D1').value = 'change of DM'
            for i in range(dm.samples.size):
                simSheet.cell('D%d' % (2 + i)).value = dm.samples[i]

            simSheet.cell('E1').value = 'change of BP'
            for i in range(bp.samples.size):
                simSheet.cell('E%d' % (2 + i)).value = bp.samples[i]

            # sheet as unhedged revenue analysis
            anaSheet = outputBook.create_sheet()
            anaSheet.title = 'ANALYSIS_UNHEDGED'
            # output mean table
            write2DTable(anaSheet, 'mean of unhedgedRevenue', kAndCListDM, kAndCListBP, unhedgedRevenueMeanList, 1)
            # output std table
            write2DTable(anaSheet, 'std of unhedgedRevenue', kAndCListDM, kAndCListBP, unhedgedRevenueStdList, 15)
            # output confident interval table
            write2DTable(anaSheet, 'min of unhedgedRevenue', kAndCListDM, kAndCListBP,
                         map(lambda x: x[0], unhedgedRevenueIntervalList), 29)
            write2DTable(anaSheet, 'max of unhedgedRevenue', kAndCListDM, kAndCListBP,
                         map(lambda x: x[1], unhedgedRevenueIntervalList), 43)
            # output propotion table
            write2DTable(anaSheet, 'proportion of unhedgedRevenue at least 706', kAndCListDM, kAndCListBP,
                         unhedgedRevenuePropList, 60)
            # output propotion confident interval table
            write2DTable(anaSheet, 'min proportion of unhedgedRevenue at least 706', kAndCListDM, kAndCListBP,
                         map(lambda x: x[0], unhedgedRevenuePropIntervalList), 74)
            write2DTable(anaSheet, 'max proportion of unhedgedRevenue at least 706', kAndCListDM, kAndCListBP,
                         map(lambda x: x[1], unhedgedRevenuePropIntervalList), 88)

        # sheet as hedged revenue analysis
        anaSheet = outputBook.create_sheet()
        anaSheet.title = 'ANALYSIS_HEDGED_nDM%d_nBP%d' % (nOfOptListDM[iOfN], nOfOptListBP[iOfN])
        # output mean table
        write2DTable(anaSheet, 'mean of hedgedRevenue', kAndCListDM, kAndCListBP, hedgedRevenueMeanList, 1)
        # output std table
        write2DTable(anaSheet, 'std of hedgedRevenue', kAndCListDM, kAndCListBP, hedgedRevenueStdList, 15)
        # output confident interval table
        write2DTable(anaSheet, 'min of hedgedRevenue', kAndCListDM, kAndCListBP,
                     map(lambda x: x[0], hedgedRevenueIntervalList), 29)
        write2DTable(anaSheet, 'max of hedgedRevenue', kAndCListDM, kAndCListBP,
                     map(lambda x: x[1], hedgedRevenueIntervalList), 43)
        # output propotion table
        write2DTable(anaSheet, 'proportion of hedgedRevenue at least 706', kAndCListDM, kAndCListBP,
                     hedgedRevenuePropList, 60)
        # output propotion confident interval table
        write2DTable(anaSheet, 'min proportion of hedgedRevenue at least 706', kAndCListDM, kAndCListBP,
                     map(lambda x: x[0], hedgedRevenuePropIntervalList), 74)
        write2DTable(anaSheet, 'max proportion of hedgedRevenue at least 706', kAndCListDM, kAndCListBP,
                     map(lambda x: x[1], hedgedRevenuePropIntervalList), 88)
    outputBook.save('output/report.xlsx')
